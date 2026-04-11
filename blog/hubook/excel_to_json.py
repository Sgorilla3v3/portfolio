#!/usr/bin/env python3
"""
청도 지역혁신가 아카이브 - 엑셀 → JSON 변환 스크립트
=====================================================
사용법:
    python excel_to_json.py                          # 기본 (같은 폴더 xlsx 자동 탐색)
    python excel_to_json.py --input 파일.xlsx        # 입력 파일 지정
    python excel_to_json.py --input 파일.xlsx --out-dir ../data  # 출력 폴더 지정
    python excel_to_json.py --vol 1                  # 특정 차수만 변환

출력:
    data/volumes.json               전체 권호 목록
    data/vol-01/meta.json           Vol.1 메타정보
    data/vol-01/innovators.json     Vol.1 혁신가 데이터
    data/vol-02/...                 (차수별 반복)
"""

import json
import re
import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

# ── 컬럼 인덱스 매핑 (0-based) ─────────────────────────────
COL = {
    "vol":          0,   # 업로드 차수
    "group":        1,   # 구분 (함께하는 기업 / 입주기업 / 함께하는 사람)
    "seq":          2,   # # (차수 내 순번)
    "name":         3,   # 이름
    "org":          4,   # 소속/직책
    "region":       6,   # 활동지역
    "domain":       7,   # 활동영역
    "activity":     8,   # 활동내용
    "projects":     9,   # 사업참여여부
    "sns":          10,  # 홈페이지 / SNS
}

# ── 차수별 표지 색상 — 흑백 계열 (순환) ───────────────────
# 완전 검정(#0d0d0d)부터 짙은 회색 계열로 미세하게 변주
COVER_COLORS = [
    "#0d0d0d", "#1a1a1a", "#262626", "#333333",
    "#111111", "#1f1f1f", "#2a2a2a", "#0a0a0a",
]

# ── SNS 라벨 → 키 매핑 ────────────────────────────────────
SNS_KEY_MAP = {
    "instagram": "instagram",
    "인스타":    "instagram",
    "인스타그램": "instagram",
    "개인":      "personal",
    "홈페이지":  "website",
    "네이버":    "naver",
    "네이버 스토어": "naver",
    "유튜브":    "youtube",
    "youtube":   "youtube",
    "facebook":  "facebook",
    "페이스북":  "facebook",
    "블로그":    "blog",
    "blog":      "blog",
    "카카오":    "kakao",
}


def parse_sns(raw: str) -> dict:
    """
    멀티라인 SNS 문자열을 {키: url} dict로 변환.
    예) "그로채 : https://...\n네이버 스토어 : naver.me/..." 
        → {"instagram": "https://...", "naver": "naver.me/..."}
    """
    if not raw:
        return {}

    result = {}
    lines = str(raw).strip().split("\n")
    for line in lines:
        line = line.strip()
        if not line:
            continue

        # "라벨 : URL" 패턴 분리
        if " : " in line:
            label, url = line.split(" : ", 1)
            label = label.strip().lower()
            url = url.strip()
        elif line.startswith("http") or line.startswith("www") or "/" in line:
            label = "website"
            url = line
        else:
            continue

        # URL 정규화 (http 없으면 추가)
        if url and not url.startswith("http"):
            url = "https://" + url

        # 라벨 → 표준 키 매핑
        key = "website"
        for keyword, mapped_key in SNS_KEY_MAP.items():
            if keyword in label or keyword in url:
                key = mapped_key
                break

        # instagram URL이면 강제 매핑
        if "instagram.com" in url:
            key = "instagram"

        # 중복 키는 numbered suffix로 구분
        if key in result:
            i = 2
            while f"{key}_{i}" in result:
                i += 1
            key = f"{key}_{i}"

        result[key] = url

    return result


def parse_domains(raw: str) -> list:
    """
    활동영역 문자열을 리스트로 변환.
    예) '"커뮤니티, 마을, 공동체", 농식품' → ["커뮤니티, 마을, 공동체", "농식품"]
    """
    if not raw:
        return []

    # 따옴표로 묶인 항목 먼저 추출
    quoted = re.findall(r'"([^"]+)"', raw)
    # 따옴표 제거 후 나머지 분리
    cleaned = re.sub(r'"[^"]+"', '', raw)
    others = [s.strip() for s in cleaned.split(',') if s.strip()]

    return quoted + others


def make_id(vol: int, seq: int) -> str:
    return f"vol{vol:02d}_{seq:03d}"


def row_to_dict(row: tuple, vol: int) -> dict | None:
    """엑셀 행 하나를 혁신가 dict로 변환. 유효하지 않으면 None 반환."""
    name = row[COL["name"]]
    if not name or str(name).strip() == "":
        return None

    seq = row[COL["seq"]] or 0

    return {
        "id":       make_id(vol, int(seq)),
        "page":     int(seq),
        "name":     str(name).strip(),
        "org":      str(row[COL["org"]] or "").strip(),
        "group":    str(row[COL["group"]] or "").strip(),   # 함께하는 기업 / 입주기업 등
        "region":   str(row[COL["region"]] or "").strip(),
        "domains":  parse_domains(str(row[COL["domain"]] or "")),
        "activity": str(row[COL["activity"]] or "").strip(),
        "projects": str(row[COL["projects"]] or "").strip(),
        "sns":      parse_sns(row[COL["sns"]]),
        # 향후 추가 예정 필드 (엑셀에 없으면 빈값)
        "tagline":  "",
        "short_bio": "",
        "story":    "",
        "thumbnail": f"images/vol-{vol:02d}/{_safe_filename(str(name))}.jpg",
    }


def _safe_filename(name: str) -> str:
    """이름을 안전한 파일명으로 변환."""
    return re.sub(r'[^\w가-힣]', '_', name).strip('_')


def load_data(xlsx_path: Path) -> dict[int, list[dict]]:
    """엑셀 전체리스트 시트에서 차수별 데이터 로드."""
    wb = load_workbook(xlsx_path, read_only=True)

    if "전체리스트" not in wb.sheetnames:
        print(f"[오류] '전체리스트' 시트를 찾을 수 없습니다.")
        print(f"  발견된 시트: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["전체리스트"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[0]  # 건너뜀

    vol_data: dict[int, list[dict]] = {}
    errors = []

    for i, row in enumerate(rows[1:], start=2):
        raw_vol = row[COL["vol"]]
        if raw_vol is None:
            continue  # 빈 행 스킵

        try:
            vol = int(raw_vol)
        except (ValueError, TypeError):
            errors.append(f"  행{i}: 업로드 차수 값 오류 ({raw_vol!r})")
            continue

        item = row_to_dict(row, vol)
        if item is None:
            continue

        if vol not in vol_data:
            vol_data[vol] = []
        vol_data[vol].append(item)

    if errors:
        print(f"[경고] 변환 중 {len(errors)}건의 오류 발생:")
        for e in errors:
            print(e)

    return vol_data


def write_vol(vol: int, innovators: list[dict], out_dir: Path, published: str = "") -> dict:
    """한 권호의 JSON 파일 생성. 권호 메타 dict 반환."""
    vol_dir = out_dir / f"vol-{vol:02d}"
    vol_dir.mkdir(parents=True, exist_ok=True)

    # 페이지 번호 순 정렬
    innovators_sorted = sorted(innovators, key=lambda x: x["page"])

    # 카테고리(domain) 수집
    all_domains = set()
    for p in innovators_sorted:
        all_domains.update(p["domains"])

    meta = {
        "vol":         vol,
        "title":       f"청도 혁신가 Vol.{vol}",
        "subtitle":    f"{vol}번째 이야기",
        "published":   published or datetime.now().strftime("%Y-%m"),
        "cover_color": COVER_COLORS[(vol - 1) % len(COVER_COLORS)],
        "cover_image": f"images/vol-{vol:02d}/cover.jpg",
        "count":       len(innovators_sorted),
        "domains":     sorted(all_domains),
        "status":      "published",
    }

    # meta.json
    (vol_dir / "meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    # innovators.json
    payload = {
        "volume":     vol,
        "innovators": innovators_sorted,
    }
    (vol_dir / "innovators.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    print(f"  [Vol.{vol}] {len(innovators_sorted)}명 → {vol_dir}/")
    return meta


def write_volumes(metas: list[dict], out_dir: Path):
    """전체 권호 목록 volumes.json 생성."""
    metas_sorted = sorted(metas, key=lambda x: x["vol"])
    payload = {"volumes": metas_sorted}
    (out_dir / "volumes.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"  [volumes.json] {len(metas_sorted)}개 권호 목록 저장")


def write_master_index(vol_data: dict[int, list[dict]], out_dir: Path):
    """전체 혁신가 통합 색인 JSON 생성 (검색용)."""
    all_innovators = []
    for vol, innovators in sorted(vol_data.items()):
        for p in innovators:
            all_innovators.append({
                "vol":       vol,
                "id":        p["id"],
                "page":      p["page"],
                "name":      p["name"],
                "org":       p["org"],
                "group":     p["group"],
                "region":    p["region"],
                "domains":   p["domains"],
                "activity":  p["activity"],   # ← 추가: 검색 히트에 필요
                "projects":  p["projects"],   # ← 추가: 검색 히트에 필요
                "tagline":   p["tagline"],
                "thumbnail": p["thumbnail"],
            })

    (out_dir / "index.json").write_text(
        json.dumps({"total": len(all_innovators), "innovators": all_innovators},
                   ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"  [index.json] 전체 {len(all_innovators)}명 통합 색인 저장")


def main():
    parser = argparse.ArgumentParser(description="청도 혁신가 엑셀 → JSON 변환기")
    parser.add_argument("--input",   "-i", type=Path, default=None,  help="입력 xlsx 경로")
    parser.add_argument("--out-dir", "-o", type=Path, default=Path("data"), help="출력 폴더 (기본: ./data)")
    parser.add_argument("--vol",     "-v", type=int,  default=None,  help="특정 차수만 변환 (기본: 전체)")
    args = parser.parse_args()

    # 입력 파일 탐색
    xlsx_path = args.input
    if xlsx_path is None:
        candidates = list(Path(".").glob("*.xlsx"))
        if not candidates:
            print("[오류] xlsx 파일을 찾을 수 없습니다. --input 옵션으로 지정하세요.")
            sys.exit(1)
        xlsx_path = candidates[0]
        print(f"[자동 감지] {xlsx_path}")

    if not xlsx_path.exists():
        print(f"[오류] 파일 없음: {xlsx_path}")
        sys.exit(1)

    out_dir: Path = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n엑셀 파일 로드 중: {xlsx_path}")
    vol_data = load_data(xlsx_path)

    if not vol_data:
        print("[오류] 변환할 데이터가 없습니다.")
        sys.exit(1)

    # 특정 차수 필터
    if args.vol is not None:
        if args.vol not in vol_data:
            print(f"[오류] Vol.{args.vol} 데이터가 없습니다. 존재하는 차수: {sorted(vol_data.keys())}")
            sys.exit(1)
        vol_data = {args.vol: vol_data[args.vol]}

    print(f"\nJSON 변환 시작 (총 {sum(len(v) for v in vol_data.values())}명, {len(vol_data)}개 권호):")
    metas = []
    for vol, innovators in sorted(vol_data.items()):
        meta = write_vol(vol, innovators, out_dir)
        metas.append(meta)

    write_volumes(metas, out_dir)
    write_master_index(vol_data, out_dir)

    print(f"\n완료! 출력 폴더: {out_dir.resolve()}")


if __name__ == "__main__":
    main()
